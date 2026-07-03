cat > netbox_manager/excel.py << 'EOF'
from pathlib import Path

from openpyxl import load_workbook


def normalise_header(value):
    if value is None:
        return ""
    return str(value).strip().lower().replace(" ", "_")


def normalise_value(value):
    if value is None:
        return ""
    return str(value).strip()


def load_sheet_rows(workbook_path, sheet_name):
    workbook_path = Path(workbook_path)

    if not workbook_path.exists():
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")

    workbook = load_workbook(workbook_path, data_only=True)

    sheet_lookup = {
        sheet.strip().lower(): sheet
        for sheet in workbook.sheetnames
    }

    key = sheet_name.strip().lower()

    if key not in sheet_lookup:
        raise ValueError(
            f"Sheet '{sheet_name}' not found. "
            f"Available sheets: {', '.join(workbook.sheetnames)}"
        )

    sheet = workbook[sheet_lookup[key]]
    rows = list(sheet.iter_rows(values_only=True))

    if not rows:
        return []

    headers = [normalise_header(value) for value in rows[0]]
    output = []

    for row_number, row in enumerate(rows[1:], start=2):
        record = {}

        for index, value in enumerate(row):
            if index >= len(headers):
                continue

            header = headers[index]

            if not header:
                continue

            record[header] = normalise_value(value)

        if not any(record.values()):
            continue

        record["_row_number"] = row_number
        output.append(record)

    return output


def require_columns(rows, required_columns, sheet_name):
    if not rows:
        return

    actual_columns = set(rows[0].keys())
    missing = required_columns - actual_columns

    if missing:
        raise ValueError(
            f"Sheet '{sheet_name}' missing columns: {', '.join(sorted(missing))}"
        )
EOF