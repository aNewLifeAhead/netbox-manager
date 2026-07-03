from pathlib import Path

from openpyxl import load_workbook


WORKBOOK_PATH = Path("data/FIX8 Network.xlsx")
WALLPLATES_SHEET = "wallplates"

DEFAULT_SITE = "FIX8Group Denton"
DEFAULT_ROLE = "Default"
DEFAULT_MANUFACTURER = "Generic"
DEFAULT_STATUS = "active"

REQUIRED_WALLPLATE_COLUMNS = {
    "name",
    "device_type",
    "location",
}


def _normalise_header(value):
    if value is None:
        return ""
    return str(value).strip().lower().replace(" ", "_")


def _normalise_value(value):
    if value is None:
        return ""
    return str(value).strip()


def _load_sheet_rows(workbook_path, sheet_name):
    if not workbook_path.exists():
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")

    workbook = load_workbook(workbook_path, data_only=True)

    sheet_lookup = {
        existing_sheet_name.strip().lower(): existing_sheet_name
        for existing_sheet_name in workbook.sheetnames
    }
    requested_sheet_key = sheet_name.strip().lower()

    if requested_sheet_key not in sheet_lookup:
        raise ValueError(
            f"Sheet '{sheet_name}' not found in {workbook_path}. "
            f"Available sheets: {', '.join(workbook.sheetnames)}"
        )

    sheet = workbook[sheet_lookup[requested_sheet_key]]
    rows = list(sheet.iter_rows(values_only=True))

    if not rows:
        return []

    headers = [_normalise_header(value) for value in rows[0]]
    output = []

    for row_number, row in enumerate(rows[1:], start=2):
        record = {}

        for index, value in enumerate(row):
            if index >= len(headers):
                continue

            header = headers[index]
            if not header:
                continue

            record[header] = _normalise_value(value)

        if not any(record.values()):
            continue

        record["_row_number"] = row_number
        output.append(record)

    return output


def _require_columns(rows, required_columns, sheet_name):
    if not rows:
        return

    actual_columns = set(rows[0].keys())
    missing_columns = required_columns - actual_columns

    if missing_columns:
        missing = ", ".join(sorted(missing_columns))
        raise ValueError(f"Sheet '{sheet_name}' is missing required columns: {missing}")


def _get_required(endpoint, label, **filters):
    obj = endpoint.get(**filters)

    if obj is None:
        rendered = ", ".join(f"{key}={value}" for key, value in filters.items())
        raise ValueError(f"{label} not found: {rendered}")

    return obj


def _get_unique_location(nb, name, site_id):
    matches = list(nb.dcim.locations.filter(name=name, site_id=site_id))

    if len(matches) == 0:
        raise ValueError(f"Location not found: {name}")

    if len(matches) > 1:
        raise ValueError(
            f"Multiple locations found named '{name}'. "
            "Make the location names unique in NetBox or the workbook."
        )

    return matches[0]


def _resolve_wallplate_values(row):
    return {
        "name": row.get("name", ""),
        "role": row.get("role") or DEFAULT_ROLE,
        "manufacturer": row.get("manufacturer") or DEFAULT_MANUFACTURER,
        "device_type": row.get("device_type", ""),
        "status": (row.get("status") or DEFAULT_STATUS).lower(),
        "site": row.get("site") or DEFAULT_SITE,
        "location": row.get("location", ""),
        "description": row.get("description", ""),
    }


def sync_wallplates(nb, dry_run=False):
    rows = _load_sheet_rows(WORKBOOK_PATH, WALLPLATES_SHEET)
    _require_columns(rows, REQUIRED_WALLPLATE_COLUMNS, WALLPLATES_SHEET)

    print(f"Reading workbook : {WORKBOOK_PATH}")
    print(f"Reading sheet    : {WALLPLATES_SHEET}")
    print(f"Rows found       : {len(rows)}")
    print()

    if not rows:
        print("No wall plate rows found. Check that your data starts below the header row.")
        return

    created = 0
    updated = 0
    unchanged = 0
    errors = 0

    for row in rows:
        row_number = row["_row_number"]
        values = _resolve_wallplate_values(row)
        name = values["name"]

        try:
            if not values["name"]:
                raise ValueError("name is required")
            if not values["location"]:
                raise ValueError("location is required")
            if not values["device_type"]:
                raise ValueError("device_type is required")

            site = _get_required(nb.dcim.sites, "Site", name=values["site"])
            role = _get_required(nb.dcim.device_roles, "Device role", name=values["role"])
            manufacturer = _get_required(
                nb.dcim.manufacturers,
                "Manufacturer",
                name=values["manufacturer"],
            )

            device_type = nb.dcim.device_types.get(
                manufacturer_id=manufacturer.id,
                model=values["device_type"],
            )

            if device_type is None:
                raise ValueError(
                    f"Device type not found: manufacturer={values['manufacturer']}, "
                    f"device_type={values['device_type']}"
                )

            location = _get_unique_location(nb, values["location"], site.id)

            payload = {
                "name": values["name"],
                "site": site.id,
                "location": location.id,
                "role": role.id,
                "device_type": device_type.id,
                "status": values["status"],
                "description": values["description"],
            }

            existing = nb.dcim.devices.get(name=values["name"], site_id=site.id)

            if existing is None:
                created += 1
                print(f"+ create {values['name']}")
                if not dry_run:
                    nb.dcim.devices.create(payload)
                continue

            desired_values = {
                "location": location.id,
                "role": role.id,
                "device_type": device_type.id,
                "status": values["status"],
                "description": values["description"],
            }

            changes = {}

            for field, desired_value in desired_values.items():
                current_value = getattr(existing, field, None)

                if hasattr(current_value, "id"):
                    current_value = current_value.id

                if current_value != desired_value:
                    changes[field] = desired_value

            if changes:
                updated += 1
                print(f"~ update {values['name']}: {', '.join(sorted(changes.keys()))}")
                if not dry_run:
                    existing.update(payload)
            else:
                unchanged += 1
                print(f"= unchanged {values['name']}")

        except Exception as error:
            errors += 1
            print(f"! row {row_number}: {name or '(blank name)'} - {error}")

    print()
    print("Wall plate sync complete")
    print(f"Created   : {created}")
    print(f"Updated   : {updated}")
    print(f"Unchanged : {unchanged}")
    print(f"Errors    : {errors}")